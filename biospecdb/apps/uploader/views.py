from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.db import transaction
from openpyxl import load_workbook

from .forms import FileUploadForm, DataInputForm
from uploader.models import Patient, Visit, SpectralData, BioSample, Symptom, Disease
from biospecdb.util import is_valid_uuid, to_uuid, num_changed_fields


def home(request):
    return render(request, 'Home.html')


@staff_member_required
def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            #return render(request, 'UploadSuccess.html')
            return display_xlsx(request)
    else:
        form = FileUploadForm()
    return render(request, 'MetadataFileUpload.html', {'form': form})


@staff_member_required
def display_xlsx(request):
    workbook = load_workbook('./biospecdb/apps/uploader/uploads/METADATA_barauna2021ultrarapid.xlsx')
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    return render(request, 'MetadataDisplay.html', {'data': data})


@staff_member_required
def data_input(request):
    message = ""
    form = DataInputForm(request=request)
    delta_count = len(form.base_fields) - 1
    
    if request.method == 'POST':
        form = DataInputForm(request.POST, request.FILES, request=request)

        if form.is_valid():
            
            changed_fields = {}
            for field_name, field_value in form.cleaned_data.items():
                # Check if the value has changed from the initial data
                initial_data = form.initial
                if field_name in form.initial and form.initial[field_name] != field_value:
                    changed_fields[field_name] = field_value
                
            if  num_changed_fields(form) < form.data.__len__(): # amount of changed fields < total amount of fields
                form.update() # Update database with changed data
                message = "Data Input with Patient ID {} has been updated successfully!!!".format(patient_id)
            else: # New entry
                form.save()  # Save data to database.
                message = "Data Input with Patient ID {} has been submitted successfully!!!".format(patient_id)
            return render(request, 'DataInputForm.html', {'form': form, 'message': message, 'delta_count': delta_count})
        
    elif request.method == 'GET':
        form = DataInputForm(request=request)
        patient_id = request.GET.get('patient_id')
        visit_date = request.GET.get('visit_date')
        if patient_id:
            if not is_valid_uuid(patient_id):
                message = "The provided Patient ID {} is not a valid number.".format(patient_id)
                return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                    'delta_count': delta_count})
            else:
                patient_id = to_uuid(patient_id)
                with transaction.atomic():
                    try:
                        patient = Patient.objects.select_for_update().get(patient_id=patient_id)
                    except (Patient.DoesNotExist):
                        message = "Data Search failed - there is no data associated with Patient ID {}." \
                            .format(patient_id)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    
                    #Analyze the situation with visits
                    previous_visits = Visit.objects.filter(patient_id=patient_id).order_by('created_at')
                    visit = None
                    if len(previous_visits) == 0:
                        message = "Data Search failed - there is no any visit of patient with Patient ID {}." \
                            .format(patient_id)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                    elif len(previous_visits) == 1:
                        visit = previous_visits.last()
                        timestamp = visit.created_at
                        visit_date = timestamp.strftime("%Y-%m-%d")
                    else: #More than one visit
                        if visit_date == '':
                            message = "There are multiple visits of the patient {} - please specify the Visit Date." \
                                .format(patient_id)
                            return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                                'delta_count': delta_count})
                        else: 
                            for current_visit in previous_visits:
                                timestamp = current_visit.created_at
                                formatted_datetime = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                                if visit_date in formatted_datetime:
                                    visit = current_visit
                                    break 
                    if visit is None:
                        message = "Data Search failed - there is no visit of patient with Patient ID {} on {}." \
                            .format(patient_id, visit_date)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                        
                    try:
                        biosample = BioSample.objects.select_for_update().get(visit=visit)
                    except (BioSample.DoesNotExist):
                        message = "Data Search failed - there is no biosample associated with the visit {}." \
                            .format(visit)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                        
                    visit_symptoms = Symptom.objects.select_for_update().filter(visit=visit)
                    symptom = visit_symptoms.order_by('days_symptomatic').last()
                    if symptom is None:
                        message = "Data Search failed - there are no symptoms associated with the visit {}." \
                            .format(visit)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                        
                    try:
                        spectraldata = SpectralData.objects.select_for_update().get(bio_sample=biosample)
                    except (SpectralData.DoesNotExist):
                        message = "Data Search failed - there is no spectral data associated with the biosample {}." \
                            .format(biosample)
                        return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                            'delta_count': delta_count})
                
                    initial_data={
                        'patient_id': patient_id,
                        'gender': patient.gender,
                        'days_symptomatic': symptom.days_symptomatic,
                        'patient_age': visit.patient_age,
                        'spectra_measurement': spectraldata.spectra_measurement,
                        'instrument': spectraldata.instrument,
                        'acquisition_time': spectraldata.acquisition_time,
                        'n_coadditions': spectraldata.n_coadditions,
                        'resolution': spectraldata.resolution,
                        'sample_type': biosample.sample_type,
                        'sample_processing': biosample.sample_processing,
                        'freezing_temp': biosample.freezing_temp,
                        'thawing_time': biosample.thawing_time,
                        'spectral_data': spectraldata.data
                    }
                    for symptom in visit_symptoms:
                        if symptom.disease.value_class == "BOOL":
                            initial_data[symptom.disease.name] = \
                                Disease.Types(symptom.disease.value_class).cast(symptom.disease_value)   
                        else:
                            initial_data[symptom.disease.name] = symptom.disease_value
                    form = DataInputForm(initial=initial_data, request=request)
                    message = "The data associated with Patient ID {} on {} is shown below:".format(patient_id, visit_date)
                    return render(request, 'DataInputForm.html', {'form': form, 'message': message, \
                        'delta_count': delta_count})
                
    else:
        form = DataInputForm(request=request)
        
    return render(request, 'DataInputForm.html', {'form': form, 'message': message, 'delta_count': delta_count})
